from openpyxl import Workbook

write_wb = Workbook()

write_ws = write_wb.create_sheet('세계지리통계')

write_ws = write_wb.active

# 좌표값 직접 입력
write_ws['A1'] = 'Geographic'

# 행 단위로 추가
write_ws.append([1,2,3])

write_ws.cell(2,2,'test')
write_wb.save('test.xlsx')